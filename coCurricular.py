import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os


def save_record(activity, name, class_):
    filename = "student_records.xlsx"
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Activity", "Name", "Class"])
        wb.save(filename)
        
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([activity, name, class_])
    wb.save(filename)
    messagebox.showinfo("Success", "Record saved successfully!")


def view_records():
    filename = "student_records.xlsx"
    if os.path.exists(filename):
        os.system(filename)
    else:
        messagebox.showerror("Error", "No records found!")


def submit_activity():
    activity = activity_var.get()
    name = name_entry.get()
    class_ = class_entry.get()
    
    if not name or not class_:
        messagebox.showerror("Error", "Please enter name and class!")
        return
    
    save_record(activity, name, class_)
    name_entry.delete(0, tk.END)
    class_entry.delete(0, tk.END)


root = tk.Tk()
root.title("Curricular Activities Registration")
root.configure(bg='lightblue')

# Creating activity selection frame
activity_frame = ttk.Frame(root)
activity_frame.pack(padx=20, pady=10)

activity_label = ttk.Label(activity_frame, text="Select Activity:")
activity_label.grid(row=0, column=0, padx=5, pady=5)

activities = ["Track & Field", "Quiz Club", "Theater & Drama", "Maths Club", 'Heritage Club','Calligraphy Club','Jewellery Making','Sudoku','Healing Club']
activity_var = tk.StringVar()
activity_combobox = ttk.Combobox(activity_frame, textvariable=activity_var, values=activities, state="readonly")
activity_combobox.grid(row=0, column=1, padx=5, pady=5)
activity_combobox.current(0)

# Creating student details entry frame
details_frame = ttk.Frame(root)
details_frame.pack(padx=20, pady=10)

name_label = ttk.Label(details_frame, text="Name:")
name_label.grid(row=0, column=0, padx=5, pady=5)
name_entry = ttk.Entry(details_frame)
name_entry.grid(row=0, column=1, padx=5, pady=5)

class_label = ttk.Label(details_frame, text="Class:")
class_label.grid(row=1, column=0, padx=5, pady=5)
class_entry = ttk.Entry(details_frame)
class_entry.grid(row=1, column=1, padx=5, pady=5)

# Submit button
submit_button = ttk.Button(root, text="Submit", command=submit_activity)
submit_button.pack(padx=20, pady=10)

# View records button
view_button = ttk.Button(root, text="View Records", command=view_records)
view_button.pack(padx=20, pady=10)

root.mainloop()
